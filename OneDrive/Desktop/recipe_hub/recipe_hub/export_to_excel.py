import os
import django
import json

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'recipe_hub.settings')
django.setup()

from recipes.models import Recipe, Category
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Create workbook
wb = Workbook()

# ==================== USERS SHEET ====================
ws_users = wb.active
ws_users.title = "Users"

# Headers
user_headers = ['ID', 'Username', 'Email', 'Is Staff', 'Is Superuser', 'Date Joined', 'Last Login']
ws_users.append(user_headers)

# Style headers
header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

for cell in ws_users[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Add user data (ordered by ID)
users = User.objects.all().order_by('id')
for user in users:
    ws_users.append([
        user.id,
        user.username,
        user.email,
        'Yes' if user.is_staff else 'No',
        'Yes' if user.is_superuser else 'No',
        user.date_joined.strftime('%Y-%m-%d %H:%M') if user.date_joined else '',
        user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never'
    ])

# Auto-adjust column widths
for column in ws_users.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_users.column_dimensions[column_letter].width = adjusted_width

# ==================== RECIPES SHEET ====================
ws_recipes = wb.create_sheet("Recipes")

# Headers
recipe_headers = ['ID', 'Title', 'Category', 'Description', 'Author', 'Prep Time (mins)', 
                  'Cook Time (mins)', 'Total Time (mins)', 'Servings', 'Created At', 'Updated At']
ws_recipes.append(recipe_headers)

# Style headers
for cell in ws_recipes[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Add recipe data (ordered by ID)
recipes = Recipe.objects.all().order_by('id')
for recipe in recipes:
    # Get category name (handle None)
    category_name = recipe.category.name if recipe.category else 'Uncategorized'
    
    ws_recipes.append([
        recipe.id,
        recipe.title,
        category_name,
        recipe.description,
        recipe.author.username,
        recipe.prep_time,
        recipe.cook_time,
        recipe.get_total_time(),
        recipe.servings,
        recipe.created_at.strftime('%Y-%m-%d %H:%M'),
        recipe.updated_at.strftime('%Y-%m-%d %H:%M')
    ])

# Auto-adjust column widths
for column in ws_recipes.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = min((max_length + 2), 50)
    ws_recipes.column_dimensions[column_letter].width = adjusted_width

# ==================== RECIPE DETAILS SHEET ====================
ws_details = wb.create_sheet("Recipe Details")

# Headers
detail_headers = ['ID', 'Title', 'Category', 'Ingredients', 'Instructions', 'Image URL']
ws_details.append(detail_headers)

# Style headers
for cell in ws_details[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Add detailed recipe data (ordered by ID)
recipes_ordered = Recipe.objects.all().order_by('id')
for recipe in recipes_ordered:
    # Get category name (handle None)
    category_name = recipe.category.name if recipe.category else 'Uncategorized'
    
    # Parse ingredients
    try:
        ingredients = json.loads(recipe.ingredients)
        ingredients_text = '\n'.join([f"• {ing}" for ing in ingredients])
    except:
        ingredients_text = recipe.ingredients
    
    ws_details.append([
        recipe.id,
        recipe.title,
        category_name,
        ingredients_text,
        recipe.instructions,
        recipe.image_url or ''
    ])

# Auto-adjust column widths and wrap text
for column in ws_details.columns:
    column_letter = column[0].column_letter
    ws_details.column_dimensions[column_letter].width = 40

# Enable text wrapping for ingredients and instructions
for row in ws_details.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# ==================== CATEGORIES SHEET ====================
ws_categories = wb.create_sheet("Categories")

# Headers
category_headers = ['ID', 'Name', 'Slug', 'Description', 'Recipe Count', 'Created At']
ws_categories.append(category_headers)

# Style headers
for cell in ws_categories[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Add category data
categories = Category.objects.all().order_by('name')
for category in categories:
    ws_categories.append([
        category.id,
        category.name,
        category.slug,
        category.description or '',
        category.recipes.count(),
        category.created_at.strftime('%Y-%m-%d %H:%M')
    ])

# Auto-adjust column widths
for column in ws_categories.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_categories.column_dimensions[column_letter].width = adjusted_width

# ==================== STATISTICS SHEET ====================
ws_stats = wb.create_sheet("Statistics")

# Add statistics
stats_data = [
    ['DATABASE STATISTICS', ''],
    ['', ''],
    ['Total Users', User.objects.count()],
    ['Admin Users', User.objects.filter(is_staff=True).count()],
    ['Regular Users', User.objects.filter(is_staff=False).count()],
    ['', ''],
    ['Total Recipes', Recipe.objects.count()],
    ['Total Categories', Category.objects.count()],
    ['', ''],
    ['RECIPES BY CATEGORY', 'COUNT'],
]

# Add category counts
for category in Category.objects.all().order_by('name'):
    count = category.recipes.count()
    stats_data.append([category.name, count])

stats_data.extend([
    ['Uncategorized', Recipe.objects.filter(category__isnull=True).count()],
    ['', ''],
    ['RECIPES BY AUTHOR', 'COUNT'],
])

# Order authors by username
for user in User.objects.all().order_by('username'):
    count = Recipe.objects.filter(author=user).count()
    if count > 0:
        stats_data.append([user.username, count])

for row in stats_data:
    ws_stats.append(row)

# Style statistics sheet
ws_stats['A1'].font = Font(size=14, bold=True)
ws_stats['A1'].fill = PatternFill(start_color='00CC66', end_color='00CC66', fill_type='solid')
ws_stats['A1'].font = Font(color='FFFFFF', size=14, bold=True)

for cell in ws_stats['A']:
    cell.font = Font(bold=True)

ws_stats.column_dimensions['A'].width = 30
ws_stats.column_dimensions['B'].width = 15

# ==================== SAVE FILE ====================
filename = "RecipeHub_Database.xlsx"

try:
    # Check if file exists
    if os.path.exists(filename):
        print("Updating existing file: " + filename)
    else:
        print("Creating new file: " + filename)

    # Save the workbook
    wb.save(filename)

    print("=" * 60)
    print("EXCEL FILE UPDATED SUCCESSFULLY!")
    print("=" * 60)
    print("Filename: " + filename)
    print("Location: " + os.path.abspath(filename))
    print("Last Updated: " + datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    print("")
    print("Current Data:")
    print("   - " + str(User.objects.count()) + " Users")
    print("   - " + str(Recipe.objects.count()) + " Recipes")
    print("   - " + str(Category.objects.count()) + " Categories")
    print("   - 5 Sheets: Users, Recipes, Recipe Details, Categories, Statistics")
    print("=" * 60)
    
except Exception as e:
    print("ERROR: " + str(e))